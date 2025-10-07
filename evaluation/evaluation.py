import os
import json
import datetime
import openpyxl
import argparse
import numpy as np
from tqdm import tqdm
from collections import defaultdict
from openpyxl.styles import PatternFill, Font


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        # print(type(v1), type(v2))
        return False
    if v1 == v2:
        return True
    else:
        return False


# BUG FIX: Color comparison bugfix for SpreadJS import/export round-trip.
# - Alpha channel is unreliable (can be 00 or FF for same color)
# - color.rgb can return various types: openpyxl objects, None, "00000000", "FF000000"
# - We normalize to "00000000" default and compare only last 6 chars (RGB values)
def _get_color_rgb(color) -> str:
    """Extract RGB value from color object, defaulting to '00000000' if not a string."""
    if color and isinstance(color.rgb, str):
        return color.rgb
    return "00000000"


def _compare_colors(color1, color2) -> bool:
    """Compare two colors using only last 6 characters (RGB), ignoring alpha channel."""
    rgb1 = _get_color_rgb(color1)
    rgb2 = _get_color_rgb(color2)
    return rgb1[-6:] == rgb2[-6:]


def compare_fill_color(fill1, fill2) -> bool:
    """Compare fill colors between two cells."""
    return _compare_colors(fill1.fgColor, fill2.fgColor) and _compare_colors(
        fill1.bgColor, fill2.bgColor
    )


def compare_font_color(font_gt, font_proc) -> bool:
    """Compare font colors between two cells."""
    return _compare_colors(font_gt.color, font_proc.color)


def col_num2name(n):
    """Convert a column number to an Excel column name"""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """Convert an Excel column name to a column number"""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    """Parse a range string like 'A1:AB12'"""
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (
        col_name2num(end_col),
        int(end_row),
    )


def generate_cell_names(range_str):
    """Generate a list of all cell names in the specified range"""
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [
        f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)
    ]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range, is_CF):
    if sheet_name not in wb_proc:
        return False, "worksheet not found"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            msg = f"Value difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.value},\
                    ws_proc has {cell_proc.value}"
            return False, msg

        if is_CF:
            if not compare_fill_color(cell_gt.fill, cell_proc.fill):
                msg = f"Fill color difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.fill.fgColor.rgb},\
                        ws_proc has {cell_proc.fill.fgColor.rgb}"
                return False, msg

            if not compare_font_color(cell_gt.font, cell_proc.font):
                # msg = f"Font color difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.font.color.rgb},\
                #        ws_proc has {cell_proc.font.color.rgb}"
                msg = f"Font color difference at cell {cell_gt.coordinate}"
                return False, msg

    print("Cell values in the specified range are identical.")
    return True, ""


def compare_workbooks(gt_file, proc_file, instruction_type, answer_position):
    if not os.path.exists(proc_file):
        return False, "File not exist"
    # Open workbooks
    if "CF" in proc_file:
        is_CF = True
    else:
        is_CF = False
    try:
        # just_open(wb_proc)
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, str(e)

    # Initialize report
    result = False
    msg = ""

    sheet_cell_ranges = answer_position.split(",")
    for sheet_cell_range in sheet_cell_ranges:
        if "!" in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split("!")
            sheet_name = sheet_name.lstrip("'").rstrip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range
    result, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range, is_CF)

    return result, msg


def parse_option():
    parser = argparse.ArgumentParser("command line arguments for evaluation.")

    parser.add_argument(
        "--master_folder",
        type=str,
        required=True,
        help="path to master folder containing spreadsheet_bench_*_run_1 subfolders",
    )
    parser.add_argument(
        "--dataset", type=str, default="all_data_912", help="dataset name"
    )

    opt = parser.parse_args()

    return opt


def evaluation(opt):
    dataset_path = os.path.abspath(f"../data/{opt.dataset}")
    with open(f"{dataset_path}/dataset.json", "r") as fp:
        dataset = json.load(fp)

    master_folder = os.path.abspath(opt.master_folder)

    eval_results = []
    missing_files = []
    total_evaluated = 0
    total_passed = 0

    for data in tqdm(dataset):
        # New file structure: spreadsheet_bench_{id}_run_1/output_ooxml.xlsx
        task_folder = f"spreadsheet_bench_{data['id']}_run_1"
        proc_path = os.path.join(master_folder, task_folder, "output_ooxml.xlsx")

        # Check if output file exists
        if not os.path.exists(proc_path):
            missing_files.append(proc_path)
            eval_results.append(
                {
                    "id": data["id"],
                    "instruction_type": data["instruction_type"],
                    "result": None,
                    "message": "Output file not found",
                }
            )
            continue

        # Get ground truth from first test case (they all share same output)
        gt_path = f"{dataset_path}/spreadsheet/{data['id']}/1_{data['id']}_answer.xlsx"

        try:
            result, message = compare_workbooks(
                gt_path,
                proc_path,
                data["instruction_type"],
                data["answer_position"],
            )
        except Exception as e:
            result = False

        total_evaluated += 1
        if result:
            total_passed += 1

        eval_results.append(
            {
                "id": data["id"],
                "instruction_type": data["instruction_type"],
                "result": int(result),
                "message": message if not result else "",
            }
        )

    # Calculate success rate (excluding missing files)
    success_rate = (total_passed / total_evaluated * 100) if total_evaluated > 0 else 0

    # Prepare summary
    summary = {
        "total_tasks": len(dataset),
        "evaluated": total_evaluated,
        "missing_files": len(missing_files),
        "passed": total_passed,
        "failed": total_evaluated - total_passed,
        "success_rate": round(success_rate, 2),
    }

    # Extract master folder name for output file
    master_folder_name = os.path.basename(master_folder)
    output_path = f"outputs/eval_{master_folder_name}.json"
    os.makedirs("outputs", exist_ok=True)

    # Save detailed results
    output_data = {
        "summary": summary,
        "missing_files": missing_files,
        "results": eval_results,
    }

    with open(output_path, "w") as fp:
        json.dump(output_data, fp, indent=4)

    # Print summary
    print(f"\n{'=' * 60}")
    print(f"Evaluation Summary")
    print(f"{'=' * 60}")
    print(f"Total tasks:       {summary['total_tasks']}")
    print(f"Evaluated:         {summary['evaluated']}")
    print(f"Missing files:     {summary['missing_files']}")
    print(f"Passed:            {summary['passed']}")
    print(f"Failed:            {summary['failed']}")
    print(f"Success rate:      {summary['success_rate']}%")
    print(f"{'=' * 60}")
    print(f"\nResults saved to {output_path}")


if __name__ == "__main__":
    opt = parse_option()
    print(opt)

    evaluation(opt)
